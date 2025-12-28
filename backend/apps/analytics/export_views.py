"""
Views para exportación de Analytics (CSV, Excel, PDF)
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Count, Q, OuterRef, Subquery, FloatField
from django.db.models.functions import TruncDate, Coalesce
from dateutil.relativedelta import relativedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from apps.turnos.models import Turno
from apps.finanzas.models import Transaction
from apps.servicios.models import Servicio
from apps.inventario.models import Producto
from apps.clientes.models import Cliente
from .permissions import IsAdminOrManager


class ExportCSVView(APIView):
    """
    GET /api/analytics/export/csv/

    Exporta datos de analytics en formato CSV
    Query params:
    - start_date: YYYY-MM-DD (requerido)
    - end_date: YYYY-MM-DD (requerido)
    """

    permission_classes = [IsAuthenticated, IsAdminOrManager]

    def get(self, request):
        # Get date range from params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            # Default to last 30 days if not provided
            end_date = datetime.now().date()
            start_date = end_date - relativedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = datetime.now().date() - relativedelta(days=30)
                end_date = datetime.now().date()

        # Get user's sucursal for multi-tenancy
        sucursal = request.user.sucursal

        # Create the HttpResponse object with CSV header
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        filename = f'analytics_{start_date}_{end_date}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Write BOM for Excel compatibility
        response.write('\ufeff')

        writer = csv.writer(response)

        # ========== SECTION 1: RESUMEN EJECUTIVO ==========
        writer.writerow(['RESUMEN EJECUTIVO'])
        writer.writerow(['Período', f'{start_date} a {end_date}'])
        writer.writerow([])

        # Get KPIs
        turnos_completados = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        )

        total_turnos = turnos_completados.count()

        # Revenue
        ingresos_totales = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Expenses
        gastos_totales = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type='EXPENSE'
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Profit
        ganancia_neta = ingresos_totales - gastos_totales

        # Clients
        clientes_nuevos = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica,
            creado_en__date__gte=start_date,
            creado_en__date__lte=end_date
        ).count()

        writer.writerow(['Métrica', 'Valor'])
        writer.writerow(['Total Ingresos', f'${ingresos_totales:,.2f}'])
        writer.writerow(['Total Gastos', f'${gastos_totales:,.2f}'])
        writer.writerow(['Ganancia Neta', f'${ganancia_neta:,.2f}'])
        writer.writerow(['Total Citas Completadas', total_turnos])
        writer.writerow(['Clientes Nuevos', clientes_nuevos])
        writer.writerow([])
        writer.writerow([])

        # ========== SECTION 2: INGRESOS POR DÍA ==========
        writer.writerow(['INGRESOS DIARIOS'])
        writer.writerow(['Fecha', 'Ingresos', 'Cantidad de Transacciones'])

        # Group income by day
        from django.db.models.functions import TruncDate

        daily_income = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).annotate(
            day=TruncDate('date')
        ).values('day').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('day')

        for item in daily_income:
            writer.writerow([
                item['day'].strftime('%Y-%m-%d'),
                f"${item['total']:,.2f}",
                item['count']
            ])

        writer.writerow([])
        writer.writerow([])

        # ========== SECTION 3: TOP SERVICIOS ==========
        writer.writerow(['TOP SERVICIOS'])
        writer.writerow(['Servicio', 'Cantidad', 'Ingresos'])

        top_services = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        ).values(
            'servicio__nombre'
        ).annotate(
            cantidad=Count('id'),
            ingresos=Sum('monto_total')
        ).order_by('-cantidad')[:20]

        for service in top_services:
            writer.writerow([
                service['servicio__nombre'] or 'Sin servicio',
                service['cantidad'],
                f"${service['ingresos'] or 0:,.2f}"
            ])

        writer.writerow([])
        writer.writerow([])

        # ========== SECTION 4: TOP PRODUCTOS ==========
        writer.writerow(['TOP PRODUCTOS'])
        writer.writerow(['Producto', 'Stock Actual', 'Precio', 'Valor Total'])

        top_products = Producto.objects.filter(
            sucursal=sucursal,
            activo=True
        ).order_by('-stock_actual')[:20]

        for product in top_products:
            valor_total = product.stock_actual * product.precio
            writer.writerow([
                product.nombre,
                product.stock_actual,
                f"${product.precio:,.2f}",
                f"${valor_total:,.2f}"
            ])

        writer.writerow([])
        writer.writerow([])

        # ========== SECTION 5: TOP CLIENTES ==========
        writer.writerow(['TOP CLIENTES'])
        writer.writerow(['Cliente', 'Email', 'Teléfono', 'Total Visitas', 'LTV'])

        # Get top clients by total spending
        from django.db.models import OuterRef, Subquery, FloatField
        from django.db.models.functions import Coalesce

        # Calculate LTV for each client
        client_spending = Transaction.objects.filter(
            client_id=OuterRef('id'),
            sucursal=sucursal,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT']
        ).values('client_id').annotate(
            total_spent=Sum('amount')
        ).values('total_spent')

        # Get clients with LTV and visit count
        top_clients = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica
        ).annotate(
            ltv=Coalesce(
                Subquery(client_spending, output_field=FloatField()),
                0.0
            ),
            visitas=Count(
                'turnos',
                filter=Q(
                    turnos__sucursal=sucursal,
                    turnos__estado='COMPLETADO'
                )
            )
        ).filter(
            visitas__gt=0
        ).order_by('-ltv')[:20]

        for client in top_clients:
            writer.writerow([
                f"{client.nombre} {client.apellido}",
                client.email or 'N/A',
                client.telefono or 'N/A',
                client.visitas,
                f"${client.ltv:,.2f}"
            ])

        writer.writerow([])
        writer.writerow([])

        # ========== SECTION 6: MÉTODOS DE PAGO ==========
        writer.writerow(['DISTRIBUCIÓN POR MÉTODO DE PAGO'])
        writer.writerow(['Método de Pago', 'Cantidad', 'Total'])

        payment_methods = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).values('payment_method').annotate(
            cantidad=Count('id'),
            total=Sum('amount')
        ).order_by('-total')

        payment_method_labels = {
            'CASH': 'Efectivo',
            'BANK_TRANSFER': 'Transferencia',
            'DEBIT_CARD': 'Débito',
            'CREDIT_CARD': 'Crédito',
            'MERCADOPAGO': 'Mercado Pago',
            'OTHER': 'Otro'
        }

        for method in payment_methods:
            label = payment_method_labels.get(method['payment_method'], method['payment_method'])
            writer.writerow([
                label,
                method['cantidad'],
                f"${method['total']:,.2f}"
            ])

        return response


class ExportExcelView(APIView):
    """
    GET /api/analytics/export/excel/

    Exporta datos de analytics en formato Excel con múltiples sheets formateadas
    Query params:
    - start_date: YYYY-MM-DD (requerido)
    - end_date: YYYY-MM-DD (requerido)
    """

    permission_classes = [IsAuthenticated, IsAdminOrManager]

    def get(self, request):
        # Get date range from params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            # Default to last 30 days if not provided
            end_date = datetime.now().date()
            start_date = end_date - relativedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = datetime.now().date() - relativedelta(days=30)
                end_date = datetime.now().date()

        # Get user's sucursal for multi-tenancy
        sucursal = request.user.sucursal

        # Create workbook
        wb = Workbook()

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== SHEET 1: RESUMEN EJECUTIVO ==========
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"

        # Title
        ws1['A1'] = 'RESUMEN EJECUTIVO - ANALYTICS'
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:D1')

        # Period
        ws1['A2'] = f'Período: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}'
        ws1['A2'].font = Font(italic=True)
        ws1.merge_cells('A2:D2')

        # Get KPIs
        turnos_completados = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        )
        total_turnos = turnos_completados.count()

        ingresos_totales = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).aggregate(total=Sum('amount'))['total'] or 0

        gastos_totales = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type='EXPENSE'
        ).aggregate(total=Sum('amount'))['total'] or 0

        ganancia_neta = ingresos_totales - gastos_totales

        clientes_nuevos = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica,
            creado_en__date__gte=start_date,
            creado_en__date__lte=end_date
        ).count()

        # Headers
        ws1['A4'] = 'Métrica'
        ws1['B4'] = 'Valor'
        for cell in ['A4', 'B4']:
            ws1[cell].fill = header_fill
            ws1[cell].font = header_font
            ws1[cell].alignment = header_alignment
            ws1[cell].border = border

        # Data
        metrics = [
            ('Total Ingresos', f'${ingresos_totales:,.2f}'),
            ('Total Gastos', f'${gastos_totales:,.2f}'),
            ('Ganancia Neta', f'${ganancia_neta:,.2f}'),
            ('Total Citas Completadas', total_turnos),
            ('Clientes Nuevos', clientes_nuevos),
        ]

        row = 5
        for metric, value in metrics:
            ws1[f'A{row}'] = metric
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
            if isinstance(value, str) and '$' in value:
                ws1[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1

        # Column widths
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20

        # ========== SHEET 2: INGRESOS DIARIOS ==========
        ws2 = wb.create_sheet("Ingresos Diarios")

        # Title
        ws2['A1'] = 'INGRESOS DIARIOS'
        ws2['A1'].font = Font(bold=True, size=14)

        # Headers
        ws2['A3'] = 'Fecha'
        ws2['B3'] = 'Ingresos'
        ws2['C3'] = 'Cantidad'
        for cell in ['A3', 'B3', 'C3']:
            ws2[cell].fill = header_fill
            ws2[cell].font = header_font
            ws2[cell].alignment = header_alignment
            ws2[cell].border = border

        # Data
        daily_income = Transaction.objects.filter(
            sucursal=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).annotate(
            day=TruncDate('date')
        ).values('day').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('day')

        row = 4
        for item in daily_income:
            ws2[f'A{row}'] = item['day'].strftime('%d/%m/%Y')
            ws2[f'B{row}'] = float(item['total'])
            ws2[f'C{row}'] = item['count']
            ws2[f'A{row}'].border = border
            ws2[f'B{row}'].border = border
            ws2[f'C{row}'].border = border
            ws2[f'B{row}'].number_format = '"$"#,##0.00'
            ws2[f'B{row}'].alignment = Alignment(horizontal='right')
            ws2[f'C{row}'].alignment = Alignment(horizontal='center')
            row += 1

        # Column widths
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 12

        # ========== SHEET 3: TOP SERVICIOS ==========
        ws3 = wb.create_sheet("Top Servicios")

        # Title
        ws3['A1'] = 'TOP SERVICIOS'
        ws3['A1'].font = Font(bold=True, size=14)

        # Headers
        ws3['A3'] = 'Servicio'
        ws3['B3'] = 'Cantidad'
        ws3['C3'] = 'Ingresos'
        for cell in ['A3', 'B3', 'C3']:
            ws3[cell].fill = header_fill
            ws3[cell].font = header_font
            ws3[cell].alignment = header_alignment
            ws3[cell].border = border

        # Data
        top_services = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        ).values(
            'servicio__nombre'
        ).annotate(
            cantidad=Count('id'),
            ingresos=Sum('monto_total')
        ).order_by('-cantidad')[:20]

        row = 4
        for service in top_services:
            ws3[f'A{row}'] = service['servicio__nombre'] or 'Sin servicio'
            ws3[f'B{row}'] = service['cantidad']
            ws3[f'C{row}'] = float(service['ingresos'] or 0)
            ws3[f'A{row}'].border = border
            ws3[f'B{row}'].border = border
            ws3[f'C{row}'].border = border
            ws3[f'B{row}'].alignment = Alignment(horizontal='center')
            ws3[f'C{row}'].number_format = '"$"#,##0.00'
            ws3[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1

        # Column widths
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 18

        # ========== SHEET 4: TOP CLIENTES ==========
        ws4 = wb.create_sheet("Top Clientes")

        # Title
        ws4['A1'] = 'TOP CLIENTES'
        ws4['A1'].font = Font(bold=True, size=14)

        # Headers
        ws4['A3'] = 'Cliente'
        ws4['B3'] = 'Email'
        ws4['C3'] = 'Teléfono'
        ws4['D3'] = 'Visitas'
        ws4['E3'] = 'LTV'
        for cell in ['A3', 'B3', 'C3', 'D3', 'E3']:
            ws4[cell].fill = header_fill
            ws4[cell].font = header_font
            ws4[cell].alignment = header_alignment
            ws4[cell].border = border

        # Data
        client_spending = Transaction.objects.filter(
            client_id=OuterRef('id'),
            sucursal=sucursal,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT']
        ).values('client_id').annotate(
            total_spent=Sum('amount')
        ).values('total_spent')

        top_clients = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica
        ).annotate(
            ltv=Coalesce(
                Subquery(client_spending, output_field=FloatField()),
                0.0
            ),
            visitas=Count(
                'turnos',
                filter=Q(
                    turnos__sucursal=sucursal,
                    turnos__estado='COMPLETADO'
                )
            )
        ).filter(
            visitas__gt=0
        ).order_by('-ltv')[:20]

        row = 4
        for client in top_clients:
            ws4[f'A{row}'] = f"{client.nombre} {client.apellido}"
            ws4[f'B{row}'] = client.email or 'N/A'
            ws4[f'C{row}'] = client.telefono or 'N/A'
            ws4[f'D{row}'] = client.visitas
            ws4[f'E{row}'] = float(client.ltv)
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws4[f'{col}{row}'].border = border
            ws4[f'D{row}'].alignment = Alignment(horizontal='center')
            ws4[f'E{row}'].number_format = '"$"#,##0.00'
            ws4[f'E{row}'].alignment = Alignment(horizontal='right')
            row += 1

        # Column widths
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 30
        ws4.column_dimensions['C'].width = 15
        ws4.column_dimensions['D'].width = 10
        ws4.column_dimensions['E'].width = 18

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'analytics_{start_date}_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class ExportPDFView(APIView):
    """
    GET /api/analytics/export/pdf/

    Exporta datos de analytics en formato PDF con tablas formateadas
    Query params:
    - start_date: YYYY-MM-DD (requerido)
    - end_date: YYYY-MM-DD (requerido)
    """

    permission_classes = [IsAuthenticated, IsAdminOrManager]

    def get(self, request):
        # Get date range from params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            # Default to last 30 days if not provided
            end_date = datetime.now().date()
            start_date = end_date - relativedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = datetime.now().date() - relativedelta(days=30)
                end_date = datetime.now().date()

        # Get user's sucursal for multi-tenancy
        sucursal = request.user.sucursal

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

        # Container for PDF elements
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=10,
            spaceBefore=15,
            fontName='Helvetica-Bold'
        )

        normal_style = styles['Normal']

        # Title
        title = Paragraph("REPORTE DE ANALYTICS", title_style)
        elements.append(title)

        # Period
        period = Paragraph(
            f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            normal_style
        )
        elements.append(period)
        elements.append(Spacer(1, 0.3*inch))

        # ========== SECTION 1: RESUMEN EJECUTIVO ==========
        elements.append(Paragraph("Resumen Ejecutivo", heading_style))

        # Get KPIs
        turnos_completados = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        )
        total_turnos = turnos_completados.count()

        ingresos_totales = Transaction.objects.filter(
            branch=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT', 'INCOME_OTHER']
        ).aggregate(total=Sum('amount'))['total'] or 0

        gastos_totales = Transaction.objects.filter(
            branch=sucursal,
            date__gte=start_date,
            date__lte=end_date,
            type='EXPENSE'
        ).aggregate(total=Sum('amount'))['total'] or 0

        ganancia_neta = ingresos_totales - gastos_totales

        clientes_nuevos = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica,
            creado_en__date__gte=start_date,
            creado_en__date__lte=end_date
        ).count()

        # KPIs table
        kpi_data = [
            ['Métrica', 'Valor'],
            ['Total Ingresos', f'${ingresos_totales:,.2f}'],
            ['Total Gastos', f'${gastos_totales:,.2f}'],
            ['Ganancia Neta', f'${ganancia_neta:,.2f}'],
            ['Total Citas Completadas', str(total_turnos)],
            ['Clientes Nuevos', str(clientes_nuevos)],
        ]

        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        elements.append(kpi_table)
        elements.append(Spacer(1, 0.3*inch))

        # ========== SECTION 2: TOP SERVICIOS ==========
        elements.append(Paragraph("Top 10 Servicios", heading_style))

        top_services = Turno.objects.filter(
            sucursal=sucursal,
            estado='COMPLETADO',
            fecha_hora_inicio__date__gte=start_date,
            fecha_hora_inicio__date__lte=end_date
        ).values(
            'servicio__nombre'
        ).annotate(
            cantidad=Count('id'),
            ingresos=Sum('monto_total')
        ).order_by('-cantidad')[:10]

        services_data = [['Servicio', 'Cantidad', 'Ingresos']]
        for service in top_services:
            services_data.append([
                service['servicio__nombre'] or 'Sin servicio',
                str(service['cantidad']),
                f"${service['ingresos'] or 0:,.2f}"
            ])

        if len(services_data) > 1:
            services_table = Table(services_data, colWidths=[3*inch, 1*inch, 1.5*inch])
            services_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

                # Body
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(services_table)
        else:
            elements.append(Paragraph("No hay datos de servicios para este período", normal_style))

        elements.append(Spacer(1, 0.3*inch))

        # ========== SECTION 3: TOP CLIENTES ==========
        elements.append(Paragraph("Top 10 Clientes", heading_style))

        # Calculate LTV for each client
        client_spending = Transaction.objects.filter(
            client_id=OuterRef('id'),
            branch=sucursal,
            type__in=['INCOME_SERVICE', 'INCOME_PRODUCT']
        ).values('client_id').annotate(
            total_spent=Sum('amount')
        ).values('total_spent')

        top_clients = Cliente.objects.filter(
            centro_estetica=sucursal.centro_estetica
        ).annotate(
            ltv=Coalesce(
                Subquery(client_spending, output_field=FloatField()),
                0.0
            ),
            visitas=Count(
                'turnos',
                filter=Q(
                    turnos__sucursal=sucursal,
                    turnos__estado='COMPLETADO'
                )
            )
        ).filter(
            visitas__gt=0
        ).order_by('-ltv')[:10]

        clients_data = [['Cliente', 'Visitas', 'LTV']]
        for client in top_clients:
            clients_data.append([
                f"{client.nombre} {client.apellido}",
                str(client.visitas),
                f"${client.ltv:,.2f}"
            ])

        if len(clients_data) > 1:
            clients_table = Table(clients_data, colWidths=[3.5*inch, 1*inch, 1.5*inch])
            clients_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

                # Body
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(clients_table)
        else:
            elements.append(Paragraph("No hay datos de clientes para este período", normal_style))

        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_text = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} - {sucursal.centro_estetica.nombre}"
        footer = Paragraph(footer_text, ParagraphStyle(
            'Footer',
            parent=normal_style,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        ))
        elements.append(footer)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(content_type='application/pdf')
        filename = f'analytics_{start_date}_{end_date}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)

        return response
